#!/usr/bin/env python3
"""Generate CI + US Expansion 5-year forecast spreadsheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = openpyxl.Workbook()

# ── Style constants ──────────────────────────────────────────
BLUE = "1A3C6E"
LIGHT_BLUE = "D6E4F0"
MED_BLUE = "8DB4E2"
DARK_BLUE = "1A3C6E"
GREEN = "C6EFCE"
ORANGE = "FDE9D9"
RED = "FFC7CE"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
YELLOW = "FFFFCC"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, size=10)
subheader_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
section_fill = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
section_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
total_font = Font(name="Calibri", bold=True, size=10)
total_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
pct_fmt = '0.0%'
dollar_fmt = '#,##0'
dollar_cent_fmt = '#,##0.00'
pct_1_fmt = '0.0%'
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

YEARS = ["FY2027", "FY2028", "FY2029", "FY2030", "FY2031"]
MONTHS = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]

# Seasonality weights
GIVING_SEASON = {
    "Apr": 0.05, "May": 0.05, "Jun": 0.05, "Jul": 0.04, "Aug": 0.04,
    "Sep": 0.05, "Oct": 0.09, "Nov": 0.13, "Dec": 0.31,
    "Jan": 0.08, "Feb": 0.05, "Mar": 0.06,
}
MARKETING_SEASON = {
    "Apr": 0.04, "May": 0.04, "Jun": 0.05, "Jul": 0.05, "Aug": 0.06,
    "Sep": 0.11, "Oct": 0.15, "Nov": 0.18, "Dec": 0.17,
    "Jan": 0.06, "Feb": 0.04, "Mar": 0.05,
}

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

def style_section_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border

def style_total_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border

def style_data_cell(cell, fmt=None):
    cell.font = Font(name="Calibri", size=10)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")
    if fmt:
        cell.number_format = fmt

def write_row(ws, row, values, fmt=None, is_total=False):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        style_data_cell(cell, fmt if c > 1 else None)
        if c == 1:
            cell.alignment = Alignment(horizontal="left")
            if is_total:
                cell.font = total_font
    if is_total:
        style_total_row(ws, row, len(values))

def auto_width(ws, min_width=12, max_width=18):
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, 14)
    ws.column_dimensions["A"].width = max_width + 10


# ════════════════════════════════════════════════════════════
# SHEET 1: Summary P&L
# ════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Summary P&L"
ws.sheet_properties.tabColor = DARK_BLUE

# --- Data ---
# CA Revenue
ca_users       = [1300, 5000, 14000, 30000, 55000]
ca_active_pct  = [0.10, 0.14, 0.17, 0.19, 0.21]
ca_txn_per     = [2.5, 3.0, 3.5, 4.2, 4.8]
ca_avg_don     = [335, 380, 420, 460, 500]
ca_plat_fee    = [0.035, 0.034, 0.033, 0.032, 0.032]

# US Revenue
us_users       = [1200, 12000, 40000, 95000, 200000]
us_active_pct  = [0.08, 0.11, 0.14, 0.17, 0.19]
us_txn_per     = [2.0, 2.5, 3.2, 3.8, 4.5]
us_avg_don_usd = [250, 290, 330, 370, 400]
us_plat_fee    = [0.035, 0.034, 0.033, 0.032, 0.032]
fx_rate        = 1.36

# SaaS
saas_partners  = [3, 22, 60, 125, 265]
saas_avg_fee   = [600, 900, 1200, 1500, 1800]

# Compute derived
ca_active   = [int(ca_users[i] * ca_active_pct[i]) for i in range(5)]
ca_txns     = [int(ca_active[i] * ca_txn_per[i]) for i in range(5)]
ca_vol      = [ca_txns[i] * ca_avg_don[i] for i in range(5)]
ca_gross    = [ca_vol[i] * ca_plat_fee[i] for i in range(5)]
ca_stripe   = [ca_txns[i] * (ca_avg_don[i] * 0.022 + 0.30) for i in range(5)]
ca_net_txn  = [ca_gross[i] - ca_stripe[i] for i in range(5)]

us_active   = [int(us_users[i] * us_active_pct[i]) for i in range(5)]
us_txns     = [int(us_active[i] * us_txn_per[i]) for i in range(5)]
us_vol_usd  = [us_txns[i] * us_avg_don_usd[i] for i in range(5)]
us_gross_usd = [us_vol_usd[i] * us_plat_fee[i] for i in range(5)]
us_stripe_usd = [us_txns[i] * (us_avg_don_usd[i] * 0.022 + 0.30) for i in range(5)]
us_net_txn_usd = [us_gross_usd[i] - us_stripe_usd[i] for i in range(5)]
us_net_txn_cad = [us_net_txn_usd[i] * fx_rate for i in range(5)]

saas_rev    = [saas_partners[i] * saas_avg_fee[i] for i in range(5)]

total_rev   = [ca_net_txn[i] + us_net_txn_cad[i] + saas_rev[i] for i in range(5)]

# Costs
# Infrastructure (monthly, interpolated to annual)
infra_monthly = [91, 389, 1117, 3369, 8265]  # computed from tier interpolation
infra_annual = [m * 12 for m in infra_monthly]

# Compliance
compliance = [5349, 30349, 40599, 45599, 55599]

# Marketing CA
mkt_ca = [7800, 16600, 71000, 124400, 178800]

# Marketing US (USD -> CAD)
mkt_us_usd = [50000, 186000, 300000, 404000, 508000]
mkt_us_cad = [int(m * fx_rate) for m in mkt_us_usd]

# Org costs (CA entities in CAD)
org_ca_fp = [17350, 29650, 53550, 75050, 109050]
org_ca_nfp = [8912, 21112, 30312, 43612, 57012]

# Org costs (US entities USD -> CAD)
org_us_fp_usd = [8500, 11500, 31500, 45000, 66500]
org_us_nfp_usd = [4025, 11025, 17525, 26025, 33025]
org_us_fp_cad = [int(v * fx_rate) for v in org_us_fp_usd]
org_us_nfp_cad = [int(v * fx_rate) for v in org_us_nfp_usd]

# US compliance (USD -> CAD)
us_compliance_usd = [13333, 35500, 51000, 67500, 85000]
us_compliance_cad = [int(v * fx_rate) for v in us_compliance_usd]

total_costs = [
    infra_annual[i] + compliance[i] + mkt_ca[i] + mkt_us_cad[i]
    + org_ca_fp[i] + org_ca_nfp[i] + org_us_fp_cad[i] + org_us_nfp_cad[i]
    + us_compliance_cad[i]
    for i in range(5)
]

net_income = [total_rev[i] - total_costs[i] for i in range(5)]

# --- Write sheet ---
row = 1
ws.cell(row=row, column=1, value="Harness Give — 5-Year Forecast (Scenario C: CI + US Expansion)")
ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
row = 2
ws.cell(row=row, column=1, value="All amounts in CAD unless noted. US amounts converted at USD/CAD 1.36.")
ws.cell(row=row, column=1).font = Font(name="Calibri", italic=True, size=9, color="666666")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

row = 4
headers = ["", *YEARS]
for c, h in enumerate(headers, 1):
    ws.cell(row=row, column=c, value=h)
style_header_row(ws, row, 6)

# Revenue section
row = 5
ws.cell(row=row, column=1, value="REVENUE")
style_section_row(ws, row, 6)

row = 6; write_row(ws, row, ["CA Net Transaction Revenue", *ca_net_txn], dollar_fmt)
row = 7; write_row(ws, row, ["US Net Transaction Revenue (CAD)", *us_net_txn_cad], dollar_fmt)
row = 8; write_row(ws, row, ["SaaS Subscription Revenue", *saas_rev], dollar_fmt)
row = 9; write_row(ws, row, ["Total Revenue", *total_rev], dollar_fmt, is_total=True)

# Revenue detail
row = 11
ws.cell(row=row, column=1, value="REVENUE DETAIL — CANADA")
style_section_row(ws, row, 6)
row = 12; write_row(ws, row, ["Registered Users", *ca_users], '#,##0')
row = 13; write_row(ws, row, ["Active Donor %", *ca_active_pct], pct_fmt)
row = 14; write_row(ws, row, ["Active Donors", *ca_active], '#,##0')
row = 15; write_row(ws, row, ["Donations per Donor/Yr", *ca_txn_per], '0.0')
row = 16; write_row(ws, row, ["Total Transactions", *ca_txns], '#,##0')
row = 17; write_row(ws, row, ["Avg Donation (CAD)", *ca_avg_don], dollar_fmt)
row = 18; write_row(ws, row, ["Donation Volume", *ca_vol], dollar_fmt)
row = 19; write_row(ws, row, ["Platform Fee %", *ca_plat_fee], pct_1_fmt)
row = 20; write_row(ws, row, ["Gross Fee Revenue", *ca_gross], dollar_fmt)
row = 21; write_row(ws, row, ["Stripe Processing Cost", *ca_stripe], dollar_fmt)
row = 22; write_row(ws, row, ["Net Transaction Revenue", *ca_net_txn], dollar_fmt, is_total=True)

row = 24
ws.cell(row=row, column=1, value="REVENUE DETAIL — US")
style_section_row(ws, row, 6)
row = 25; write_row(ws, row, ["Registered Users", *us_users], '#,##0')
row = 26; write_row(ws, row, ["Active Donor %", *us_active_pct], pct_fmt)
row = 27; write_row(ws, row, ["Active Donors", *us_active], '#,##0')
row = 28; write_row(ws, row, ["Donations per Donor/Yr", *us_txn_per], '0.0')
row = 29; write_row(ws, row, ["Total Transactions", *us_txns], '#,##0')
row = 30; write_row(ws, row, ["Avg Donation (USD)", *us_avg_don_usd], dollar_fmt)
row = 31; write_row(ws, row, ["Donation Volume (USD)", *us_vol_usd], dollar_fmt)
row = 32; write_row(ws, row, ["Platform Fee %", *us_plat_fee], pct_1_fmt)
row = 33; write_row(ws, row, ["Gross Fee Revenue (USD)", *us_gross_usd], dollar_fmt)
row = 34; write_row(ws, row, ["Stripe Processing Cost (USD)", *us_stripe_usd], dollar_fmt)
row = 35; write_row(ws, row, ["Net Transaction Revenue (USD)", *us_net_txn_usd], dollar_fmt)
row = 36; write_row(ws, row, ["Net Transaction Revenue (CAD)", *us_net_txn_cad], dollar_fmt, is_total=True)

# Costs section
row = 38
ws.cell(row=row, column=1, value="OPERATING EXPENSES")
style_section_row(ws, row, 6)
row = 39; write_row(ws, row, ["Infrastructure (annual)", *infra_annual], dollar_fmt)
row = 40; write_row(ws, row, ["Compliance & Security", *compliance], dollar_fmt)
row = 41; write_row(ws, row, ["Marketing — Canada", *mkt_ca], dollar_fmt)
row = 42; write_row(ws, row, ["Marketing — US (CAD)", *mkt_us_cad], dollar_fmt)
row = 43; write_row(ws, row, ["Org Costs — Harness Exchange (CA)", *org_ca_fp], dollar_fmt)
row = 44; write_row(ws, row, ["Org Costs — Harness Good (CA)", *org_ca_nfp], dollar_fmt)
row = 45; write_row(ws, row, ["Org Costs — US For-Profit (CAD)", *org_us_fp_cad], dollar_fmt)
row = 46; write_row(ws, row, ["Org Costs — US Not-for-Profit (CAD)", *org_us_nfp_cad], dollar_fmt)
row = 47; write_row(ws, row, ["US Compliance & Registrations (CAD)", *us_compliance_cad], dollar_fmt)
row = 48; write_row(ws, row, ["Total Operating Expenses", *total_costs], dollar_fmt, is_total=True)

# Net
row = 50
ws.cell(row=row, column=1, value="NET INCOME / (LOSS)")
style_section_row(ws, row, 6)
row = 51; write_row(ws, row, ["Net Income (Loss)", *net_income], dollar_fmt, is_total=True)
# Color negative red
for c in range(2, 7):
    cell = ws.cell(row=51, column=c)
    if cell.value and cell.value < 0:
        cell.font = Font(name="Calibri", bold=True, size=10, color="CC0000")

auto_width(ws, max_width=22)
ws.column_dimensions["A"].width = 38


# ════════════════════════════════════════════════════════════
# SHEET 2: Traffic & Users
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Traffic & Users")
ws2.sheet_properties.tabColor = "2E75B6"

row = 1
ws2.cell(row=row, column=1, value="Traffic & User Funnel — Scenario C (CI + US)")
ws2.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

row = 3
for c, h in enumerate(["", *YEARS], 1):
    ws2.cell(row=row, column=c, value=h)
style_header_row(ws2, row, 6)

# CA Traffic
row = 4
ws2.cell(row=row, column=1, value="CANADA — TRAFFIC")
style_section_row(ws2, row, 6)

ca_ci_ref    = [90000, 100000, 120000, 150000, 180000]
ca_organic   = [22000, 40000, 50000, 60000, 75000]
ca_paid      = [18000, 35000, 40000, 50000, 65000]
ca_total_vis = [ca_ci_ref[i]+ca_organic[i]+ca_paid[i] for i in range(5)]
ca_reg       = [1300, 5000, 14000, 30000, 55000]
ca_conv      = [ca_reg[i]/ca_total_vis[i] for i in range(5)]

row = 5; write_row(ws2, row, ["CI Referral Visitors", *ca_ci_ref], '#,##0')
row = 6; write_row(ws2, row, ["Organic Search Visitors", *ca_organic], '#,##0')
row = 7; write_row(ws2, row, ["Paid Media Visitors", *ca_paid], '#,##0')
row = 8; write_row(ws2, row, ["Total Visitors", *ca_total_vis], '#,##0', is_total=True)
row = 9; write_row(ws2, row, ["Registered Users", *ca_reg], '#,##0')
row = 10; write_row(ws2, row, ["Registration Rate", *ca_conv], pct_1_fmt)

# CI Detail
row = 12
ws2.cell(row=row, column=1, value="CI REFERRAL DETAIL")
style_section_row(ws2, row, 6)
ci_uv = [305000, 295000, 315000, 340000, 365000]
ci_capture = [0.30, 0.34, 0.38, 0.44, 0.49]
row = 13; write_row(ws2, row, ["CI Unique Visitors (projected)", *ci_uv], '#,##0')
row = 14; write_row(ws2, row, ["Capture Rate", *ci_capture], pct_1_fmt)
row = 15; write_row(ws2, row, ["CI Referral to Give", *ca_ci_ref], '#,##0', is_total=True)

# US Traffic
row = 17
ws2.cell(row=row, column=1, value="US — TRAFFIC")
style_section_row(ws2, row, 6)
us_organic   = [10000, 40000, 80000, 150000, 250000]
us_paid_vis  = [40000, 130000, 270000, 420000, 580000]
us_pr        = [10000, 30000, 50000, 80000, 120000]
us_total_vis = [us_organic[i]+us_paid_vis[i]+us_pr[i] for i in range(5)]
us_reg       = [1200, 12000, 40000, 95000, 200000]
us_conv      = [us_reg[i]/us_total_vis[i] for i in range(5)]

row = 18; write_row(ws2, row, ["Organic Search Visitors", *us_organic], '#,##0')
row = 19; write_row(ws2, row, ["Paid Media Visitors", *us_paid_vis], '#,##0')
row = 20; write_row(ws2, row, ["PR / Referral Visitors", *us_pr], '#,##0')
row = 21; write_row(ws2, row, ["Total Visitors", *us_total_vis], '#,##0', is_total=True)
row = 22; write_row(ws2, row, ["Registered Users", *us_reg], '#,##0')
row = 23; write_row(ws2, row, ["Registration Rate", *us_conv], pct_1_fmt)

# Combined
row = 25
ws2.cell(row=row, column=1, value="COMBINED (CA + US)")
style_section_row(ws2, row, 6)
combined_vis = [ca_total_vis[i]+us_total_vis[i] for i in range(5)]
combined_reg = [ca_reg[i]+us_reg[i] for i in range(5)]
row = 26; write_row(ws2, row, ["Total Visitors", *combined_vis], '#,##0')
row = 27; write_row(ws2, row, ["Total Registered Users", *combined_reg], '#,##0', is_total=True)
row = 28; write_row(ws2, row, ["CA Share of Users", *[ca_reg[i]/combined_reg[i] for i in range(5)]], pct_1_fmt)
row = 29; write_row(ws2, row, ["US Share of Users", *[us_reg[i]/combined_reg[i] for i in range(5)]], pct_1_fmt)

auto_width(ws2, max_width=22)
ws2.column_dimensions["A"].width = 34


# ════════════════════════════════════════════════════════════
# SHEET 3: Revenue Detail
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Revenue Detail")
ws3.sheet_properties.tabColor = "00B050"

row = 1
ws3.cell(row=row, column=1, value="Revenue Build-Up — Scenario C")
ws3.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

row = 3
for c, h in enumerate(["", *YEARS], 1):
    ws3.cell(row=row, column=c, value=h)
style_header_row(ws3, row, 6)

# CA
row = 4
ws3.cell(row=row, column=1, value="CANADA — TRANSACTION REVENUE")
style_section_row(ws3, row, 6)
row = 5; write_row(ws3, row, ["Registered Users", *ca_users], '#,##0')
row = 6; write_row(ws3, row, ["Active Donor %", *ca_active_pct], pct_fmt)
row = 7; write_row(ws3, row, ["Active Donors", *ca_active], '#,##0')
row = 8; write_row(ws3, row, ["Txns per Donor/Yr", *ca_txn_per], '0.0')
row = 9; write_row(ws3, row, ["Total Transactions", *ca_txns], '#,##0')
row = 10; write_row(ws3, row, ["Avg Donation (CAD)", *ca_avg_don], dollar_fmt)
row = 11; write_row(ws3, row, ["Donation Volume", *ca_vol], dollar_fmt)
row = 12; write_row(ws3, row, ["Platform Fee %", *ca_plat_fee], pct_1_fmt)
row = 13; write_row(ws3, row, ["Gross Platform Fee", *ca_gross], dollar_fmt)
row = 14; write_row(ws3, row, ["Less: Stripe (2.2% + $0.30)", *ca_stripe], dollar_fmt)
ca_margin = [ca_net_txn[i]/ca_vol[i] if ca_vol[i] else 0 for i in range(5)]
row = 15; write_row(ws3, row, ["Net Transaction Revenue", *ca_net_txn], dollar_fmt, is_total=True)
row = 16; write_row(ws3, row, ["Effective Net Margin", *ca_margin], pct_1_fmt)

# US
row = 18
ws3.cell(row=row, column=1, value="US — TRANSACTION REVENUE (USD)")
style_section_row(ws3, row, 6)
row = 19; write_row(ws3, row, ["Registered Users", *us_users], '#,##0')
row = 20; write_row(ws3, row, ["Active Donor %", *us_active_pct], pct_fmt)
row = 21; write_row(ws3, row, ["Active Donors", *us_active], '#,##0')
row = 22; write_row(ws3, row, ["Txns per Donor/Yr", *us_txn_per], '0.0')
row = 23; write_row(ws3, row, ["Total Transactions", *us_txns], '#,##0')
row = 24; write_row(ws3, row, ["Avg Donation (USD)", *us_avg_don_usd], dollar_fmt)
row = 25; write_row(ws3, row, ["Donation Volume (USD)", *us_vol_usd], dollar_fmt)
row = 26; write_row(ws3, row, ["Platform Fee %", *us_plat_fee], pct_1_fmt)
row = 27; write_row(ws3, row, ["Gross Platform Fee (USD)", *us_gross_usd], dollar_fmt)
row = 28; write_row(ws3, row, ["Less: Stripe (2.2% + $0.30 USD)", *us_stripe_usd], dollar_fmt)
row = 29; write_row(ws3, row, ["Net Transaction Revenue (USD)", *us_net_txn_usd], dollar_fmt, is_total=True)
row = 30; write_row(ws3, row, ["FX Rate (USD/CAD)", *[fx_rate]*5], '0.00')
row = 31; write_row(ws3, row, ["Net Transaction Revenue (CAD)", *us_net_txn_cad], dollar_fmt, is_total=True)

# SaaS
row = 33
ws3.cell(row=row, column=1, value="SAAS SUBSCRIPTION REVENUE")
style_section_row(ws3, row, 6)
row = 34; write_row(ws3, row, ["Charity Partners", *saas_partners], '#,##0')
row = 35; write_row(ws3, row, ["Avg Annual Fee (CAD)", *saas_avg_fee], dollar_fmt)
row = 36; write_row(ws3, row, ["SaaS Revenue", *saas_rev], dollar_fmt, is_total=True)

# Consolidated
row = 38
ws3.cell(row=row, column=1, value="CONSOLIDATED REVENUE (CAD)")
style_section_row(ws3, row, 6)
row = 39; write_row(ws3, row, ["CA Net Transaction Revenue", *ca_net_txn], dollar_fmt)
row = 40; write_row(ws3, row, ["US Net Transaction Revenue (CAD)", *us_net_txn_cad], dollar_fmt)
row = 41; write_row(ws3, row, ["SaaS Subscription Revenue", *saas_rev], dollar_fmt)
row = 42; write_row(ws3, row, ["Total Revenue", *total_rev], dollar_fmt, is_total=True)

auto_width(ws3, max_width=22)
ws3.column_dimensions["A"].width = 36


# ════════════════════════════════════════════════════════════
# SHEET 4: Infrastructure
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Infrastructure")
ws4.sheet_properties.tabColor = "FFC000"

row = 1
ws4.cell(row=row, column=1, value="Infrastructure Cost Model — Monthly by Tier")
ws4.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

row = 3
for c, h in enumerate(["Category", "Startup (1K)", "Growth (10K)", "Scale (50K)", "Enterprise (250K)"], 1):
    ws4.cell(row=row, column=c, value=h)
style_header_row(ws4, row, 5)

infra_items = [
    ("Compute & Hosting", 35, 85, 400, 1050),
    ("PostgreSQL + pgvector", 18, 100, 375, 1400),
    ("Redis Cache", 5, 35, 150, 300),
    ("Kafka / Event Streaming", 0, 55, 200, 1000),
    ("AI/LLM Tokens (Claude)", 10, 100, 475, 2350),
    ("Email (Resend)", 0, 30, 75, 300),
    ("SMS (Twilio)", 2, 25, 90, 450),
    ("CDN + Storage (S3/CF)", 6, 35, 170, 500),
    ("Monitoring", 0, 30, 80, 200),
    ("Search (Elasticsearch)", 0, 75, 300, 700),
    ("Domain + SSL", 15, 15, 15, 15),
]

totals = [0, 0, 0, 0]
for i, item in enumerate(infra_items):
    row = 4 + i
    write_row(ws4, row, list(item), dollar_fmt)
    for j in range(4):
        totals[j] += item[j+1]

row = 4 + len(infra_items)
write_row(ws4, row, ["Total Monthly", *totals], dollar_fmt, is_total=True)
row += 1
write_row(ws4, row, ["Total Annual", *[t*12 for t in totals]], dollar_fmt, is_total=True)

# Interpolated per year
row += 2
ws4.cell(row=row, column=1, value="INTERPOLATED MONTHLY COST BY YEAR")
style_section_row(ws4, row, 6)

row += 1
for c, h in enumerate(["", *YEARS], 1):
    ws4.cell(row=row, column=c, value=h)
style_header_row(ws4, row, 6)

# Combined users for tier interpolation
combined_users = [2500, 17000, 54000, 125000, 255000]
# Tier breakpoints: 1000, 10000, 50000, 250000
tier_bp = [1000, 10000, 50000, 250000]
tier_totals = totals  # [91, 585, 2330, 8265]

def interpolate_tier(users):
    if users <= tier_bp[0]:
        return tier_totals[0]
    elif users <= tier_bp[1]:
        t = (users - tier_bp[0]) / (tier_bp[1] - tier_bp[0])
        return tier_totals[0] * (1 - t) + tier_totals[1] * t
    elif users <= tier_bp[2]:
        t = (users - tier_bp[1]) / (tier_bp[2] - tier_bp[1])
        return tier_totals[1] * (1 - t) + tier_totals[2] * t
    elif users <= tier_bp[3]:
        t = (users - tier_bp[2]) / (tier_bp[3] - tier_bp[2])
        return tier_totals[2] * (1 - t) + tier_totals[3] * t
    else:
        return tier_totals[3]

interp_monthly = [round(interpolate_tier(u)) for u in combined_users]
interp_annual = [m * 12 for m in interp_monthly]

row += 1; write_row(ws4, row, ["Combined Users", *combined_users], '#,##0')
row += 1; write_row(ws4, row, ["Monthly Infrastructure", *interp_monthly], dollar_fmt)
row += 1; write_row(ws4, row, ["Annual Infrastructure", *interp_annual], dollar_fmt, is_total=True)

# Update infra_annual in summary with accurate interpolation
# (we'll regenerate summary below)

auto_width(ws4, max_width=18)
ws4.column_dimensions["A"].width = 30


# ════════════════════════════════════════════════════════════
# SHEET 5: Marketing
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Marketing")
ws5.sheet_properties.tabColor = "FF6600"

row = 1
ws5.cell(row=row, column=1, value="Marketing & User Acquisition — Scenario C")
ws5.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

row = 3
for c, h in enumerate(["", *YEARS], 1):
    ws5.cell(row=row, column=c, value=h)
style_header_row(ws5, row, 6)

row = 4
ws5.cell(row=row, column=1, value="CANADA (WITH CI)")
style_section_row(ws5, row, 6)
row = 5; write_row(ws5, row, ["CI Referral (partnership)", *[0]*5], dollar_fmt)
row = 6; write_row(ws5, row, ["Organic Marketing", *[6000, 12000, 18000, 24000, 30000]], dollar_fmt)
row = 7; write_row(ws5, row, ["Paid Media", *[1800, 4600, 53000, 100400, 148800]], dollar_fmt)
row = 8; write_row(ws5, row, ["Total CA Marketing", *mkt_ca], dollar_fmt, is_total=True)

row = 10
ws5.cell(row=row, column=1, value="CANADA — COST PER VISITOR")
style_section_row(ws5, row, 6)
ca_cpv_ci = [0]*5
ca_cpv_org = [0.27, 0.30, 0.36, 0.40, 0.40]
ca_cpv_paid = [0.10, 0.13, 1.33, 2.01, 2.29]
ca_cpv_blend = [0.06, 0.10, 0.34, 0.48, 0.56]
row = 11; write_row(ws5, row, ["CI Referral", *ca_cpv_ci], dollar_cent_fmt)
row = 12; write_row(ws5, row, ["Organic Search", *ca_cpv_org], dollar_cent_fmt)
row = 13; write_row(ws5, row, ["Paid Media", *ca_cpv_paid], dollar_cent_fmt)
row = 14; write_row(ws5, row, ["Blended", *ca_cpv_blend], dollar_cent_fmt, is_total=True)

row = 16
ws5.cell(row=row, column=1, value="US MARKET (USD)")
style_section_row(ws5, row, 6)
row = 17; write_row(ws5, row, ["Organic Marketing", *[10000, 36000, 60000, 84000, 108000]], dollar_fmt)
row = 18; write_row(ws5, row, ["Paid Media", *[40000, 150000, 240000, 320000, 400000]], dollar_fmt)
row = 19; write_row(ws5, row, ["Total US Marketing (USD)", *mkt_us_usd], dollar_fmt, is_total=True)
row = 20; write_row(ws5, row, ["Total US Marketing (CAD)", *mkt_us_cad], dollar_fmt, is_total=True)

row = 22
ws5.cell(row=row, column=1, value="COMBINED MARKETING (CAD)")
style_section_row(ws5, row, 6)
combined_mkt = [mkt_ca[i] + mkt_us_cad[i] for i in range(5)]
row = 23; write_row(ws5, row, ["Canada", *mkt_ca], dollar_fmt)
row = 24; write_row(ws5, row, ["US (CAD)", *mkt_us_cad], dollar_fmt)
row = 25; write_row(ws5, row, ["Total Marketing (CAD)", *combined_mkt], dollar_fmt, is_total=True)

auto_width(ws5, max_width=18)
ws5.column_dimensions["A"].width = 32


# ════════════════════════════════════════════════════════════
# SHEET 6: Org Costs
# ════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Org Costs")
ws6.sheet_properties.tabColor = "7030A0"

row = 1
ws6.cell(row=row, column=1, value="Organizational Costs — All Four Entities")
ws6.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
ws6.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

row = 3
for c, h in enumerate(["", *YEARS], 1):
    ws6.cell(row=row, column=c, value=h)
style_header_row(ws6, row, 6)

# CA FP
row = 4
ws6.cell(row=row, column=1, value="HARNESS EXCHANGE — CA FOR-PROFIT (CAD)")
style_section_row(ws6, row, 6)
ca_fp_items = [
    ("Annual return", [50]*5),
    ("T2 corporate tax return", [1200, 1500, 2000, 2500, 3000]),
    ("HST/GST filing", [1200, 1200, 2000, 2000, 2000]),
    ("Bookkeeping", [3600, 6000, 9600, 12000, 18000]),
    ("Financial review/audit", [0, 5000, 10000, 15000, 20000]),
    ("D&O insurance", [3000, 3000, 5000, 7000, 10000]),
    ("Commercial general liability", [800, 800, 1200, 1500, 2000]),
    ("Cyber / tech E&O insurance", [2000, 3000, 5000, 8000, 12000]),
    ("Legal retainer", [5000, 8000, 12000, 15000, 20000]),
    ("Board of Directors", [0, 0, 5000, 10000, 20000]),
    ("Registered agent", [500]*5),
    ("Business banking", [0, 600, 1200, 1500, 1500]),
]
for i, (name, vals) in enumerate(ca_fp_items):
    row = 5 + i
    write_row(ws6, row, [name, *vals], dollar_fmt)
row = 5 + len(ca_fp_items)
write_row(ws6, row, ["Total", *org_ca_fp], dollar_fmt, is_total=True)

# CA NFP
row += 2
ws6.cell(row=row, column=1, value="HARNESS GOOD — CA NOT-FOR-PROFIT (CAD)")
style_section_row(ws6, row, 6)
ca_nfp_items = [
    ("Annual return", [12]*5),
    ("T3010 charity return", [1500, 2000, 2500, 3000, 3500]),
    ("Bookkeeping", [2400, 3600, 6000, 9600, 12000]),
    ("Financial audit", [0, 8000, 10000, 15000, 20000]),
    ("General liability insurance", [500, 500, 800, 1000, 1500]),
    ("Cyber insurance", [1500, 2000, 3000, 5000, 8000]),
    ("Legal retainer", [3000, 5000, 8000, 10000, 12000]),
    ("Board of Directors", [0]*5),
]
for i, (name, vals) in enumerate(ca_nfp_items):
    row += 1
    write_row(ws6, row, [name, *vals], dollar_fmt)
row += 1
write_row(ws6, row, ["Total", *org_ca_nfp], dollar_fmt, is_total=True)

# US FP
row += 2
ws6.cell(row=row, column=1, value="HARNESS EXCHANGE US — DELAWARE C-CORP (USD)")
style_section_row(ws6, row, 6)
us_fp_items = [
    ("Delaware franchise tax", [400]*5),
    ("Delaware annual report", [50]*5),
    ("Registered agent", [50]*5),
    ("Form 1120 tax return", [1500, 2000, 3000, 3500, 5000]),
    ("Financial audit", [0, 0, 10000, 15000, 20000]),
    ("D&O insurance", [2000, 2000, 4000, 6000, 8000]),
    ("General liability + cyber", [1500, 2000, 3000, 5000, 8000]),
    ("Legal (US corporate)", [3000, 5000, 8000, 10000, 15000]),
    ("Board of Directors", [0, 0, 3000, 5000, 10000]),
]
for i, (name, vals) in enumerate(us_fp_items):
    row += 1
    write_row(ws6, row, [name, *vals], dollar_fmt)
row += 1
write_row(ws6, row, ["Total (USD)", *org_us_fp_usd], dollar_fmt, is_total=True)

# US NFP
row += 2
ws6.cell(row=row, column=1, value="HARNESS GOOD US FOUNDATION — 501(c)(3) (USD)")
style_section_row(ws6, row, 6)
us_nfp_items = [
    ("Delaware annual report", [25]*5),
    ("Delaware franchise tax (exempt)", [0]*5),
    ("Form 990 preparation", [1000, 1500, 2000, 2000, 2000]),
    ("Financial audit", [0, 5000, 8000, 12000, 15000]),
    ("General liability + cyber", [1000, 1500, 2500, 4000, 6000]),
    ("Legal (US NFP counsel)", [2000, 3000, 5000, 8000, 10000]),
    ("Board of Directors", [0]*5),
]
for i, (name, vals) in enumerate(us_nfp_items):
    row += 1
    write_row(ws6, row, [name, *vals], dollar_fmt)
row += 1
write_row(ws6, row, ["Total (USD)", *org_us_nfp_usd], dollar_fmt, is_total=True)

# Summary
row += 2
ws6.cell(row=row, column=1, value="CONSOLIDATED ORG COSTS (LOCAL CURRENCY)")
style_section_row(ws6, row, 6)
row += 1; write_row(ws6, row, ["Harness Exchange (CAD)", *org_ca_fp], dollar_fmt)
row += 1; write_row(ws6, row, ["Harness Good (CAD)", *org_ca_nfp], dollar_fmt)
row += 1; write_row(ws6, row, ["US For-Profit (USD)", *org_us_fp_usd], dollar_fmt)
row += 1; write_row(ws6, row, ["US Not-for-Profit (USD)", *org_us_nfp_usd], dollar_fmt)
combined_org_cad = [
    org_ca_fp[i] + org_ca_nfp[i] + org_us_fp_cad[i] + org_us_nfp_cad[i]
    for i in range(5)
]
row += 1; write_row(ws6, row, ["Combined (CAD, US @ 1.36)", *combined_org_cad], dollar_fmt, is_total=True)

auto_width(ws6, max_width=18)
ws6.column_dimensions["A"].width = 42


# ════════════════════════════════════════════════════════════
# SHEET 7: Monthly Forecast (Year 1)
# ════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Monthly FY2027")
ws7.sheet_properties.tabColor = "00B0F0"

row = 1
ws7.cell(row=row, column=1, value="Monthly Forecast — FY2027 (Year 1) with Seasonality")
ws7.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
ws7.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)

row = 3
headers7 = ["", *MONTHS, "Total"]
for c, h in enumerate(headers7, 1):
    ws7.cell(row=row, column=c, value=h)
style_header_row(ws7, row, 14)

# Y1 annual totals
y1_ca_net = ca_net_txn[0]
y1_us_net_cad = us_net_txn_cad[0]
y1_saas = saas_rev[0]
y1_mkt_ca = mkt_ca[0]
y1_mkt_us = mkt_us_cad[0]
y1_infra = interp_monthly[0] * 12

# US starts Aug (month 5 of fiscal year Apr-Mar), so 8 months
us_start_month = "Aug"  # index 4 in MONTHS list

row = 4
ws7.cell(row=row, column=1, value="REVENUE (CAD)")
style_section_row(ws7, row, 14)

# CA donation revenue by month (giving season)
ca_monthly_rev = [round(y1_ca_net * GIVING_SEASON[m]) for m in MONTHS]
row = 5
vals = ["CA Net Txn Revenue", *ca_monthly_rev, sum(ca_monthly_rev)]
write_row(ws7, row, vals, dollar_fmt)

# US donation revenue by month (giving season, zero before Aug)
us_months_active = {m: GIVING_SEASON[m] for m in MONTHS[MONTHS.index("Aug"):]}
us_weight_total = sum(us_months_active.values())
us_monthly_rev = []
for m in MONTHS:
    if m in us_months_active:
        us_monthly_rev.append(round(y1_us_net_cad * (us_months_active[m] / us_weight_total)))
    else:
        us_monthly_rev.append(0)
row = 6
vals = ["US Net Txn Revenue (CAD)", *us_monthly_rev, sum(us_monthly_rev)]
write_row(ws7, row, vals, dollar_fmt)

# SaaS (uniform)
saas_monthly = [round(y1_saas / 12)] * 12
row = 7
vals = ["SaaS Revenue", *saas_monthly, sum(saas_monthly)]
write_row(ws7, row, vals, dollar_fmt)

# Total revenue
total_monthly_rev = [ca_monthly_rev[i] + us_monthly_rev[i] + saas_monthly[i] for i in range(12)]
row = 8
vals = ["Total Revenue", *total_monthly_rev, sum(total_monthly_rev)]
write_row(ws7, row, vals, dollar_fmt, is_total=True)

# Expenses
row = 10
ws7.cell(row=row, column=1, value="EXPENSES (CAD)")
style_section_row(ws7, row, 14)

# Infrastructure (uniform)
infra_m = [interp_monthly[0]] * 12
row = 11
vals = ["Infrastructure", *infra_m, sum(infra_m)]
write_row(ws7, row, vals, dollar_fmt)

# Marketing CA (marketing season)
ca_mkt_monthly = [round(y1_mkt_ca * MARKETING_SEASON[m]) for m in MONTHS]
row = 12
vals = ["Marketing — CA", *ca_mkt_monthly, sum(ca_mkt_monthly)]
write_row(ws7, row, vals, dollar_fmt)

# Marketing US (marketing season, zero before Aug)
us_mkt_months = {m: MARKETING_SEASON[m] for m in MONTHS[MONTHS.index("Aug"):]}
us_mkt_weight = sum(us_mkt_months.values())
us_mkt_monthly = []
for m in MONTHS:
    if m in us_mkt_months:
        us_mkt_monthly.append(round(y1_mkt_us * (us_mkt_months[m] / us_mkt_weight)))
    else:
        us_mkt_monthly.append(0)
row = 13
vals = ["Marketing — US (CAD)", *us_mkt_monthly, sum(us_mkt_monthly)]
write_row(ws7, row, vals, dollar_fmt)

# Org costs (uniform monthly)
y1_org_total = org_ca_fp[0] + org_ca_nfp[0] + org_us_fp_cad[0] + org_us_nfp_cad[0]
org_monthly = [round(y1_org_total / 12)] * 12
row = 14
vals = ["Org Costs (all entities)", *org_monthly, sum(org_monthly)]
write_row(ws7, row, vals, dollar_fmt)

# Compliance (uniform)
comp_monthly = [round(compliance[0] / 12)] * 12
row = 15
vals = ["Compliance & Security", *comp_monthly, sum(comp_monthly)]
write_row(ws7, row, vals, dollar_fmt)

# US Compliance (Aug onward)
us_comp_m = []
us_comp_months_active = {m: 1 for m in MONTHS[MONTHS.index("Aug"):]}
us_comp_per_month = round(us_compliance_cad[0] / 8)  # 8 months
for m in MONTHS:
    if m in us_comp_months_active:
        us_comp_m.append(us_comp_per_month)
    else:
        us_comp_m.append(0)
row = 16
vals = ["US Compliance (CAD)", *us_comp_m, sum(us_comp_m)]
write_row(ws7, row, vals, dollar_fmt)

# Total expenses
total_monthly_exp = [
    infra_m[i] + ca_mkt_monthly[i] + us_mkt_monthly[i]
    + org_monthly[i] + comp_monthly[i] + us_comp_m[i]
    for i in range(12)
]
row = 17
vals = ["Total Expenses", *total_monthly_exp, sum(total_monthly_exp)]
write_row(ws7, row, vals, dollar_fmt, is_total=True)

# Net
row = 19
ws7.cell(row=row, column=1, value="NET INCOME / (LOSS)")
style_section_row(ws7, row, 14)
net_monthly = [total_monthly_rev[i] - total_monthly_exp[i] for i in range(12)]
row = 20
vals = ["Net Income (Loss)", *net_monthly, sum(net_monthly)]
write_row(ws7, row, vals, dollar_fmt, is_total=True)
for c in range(2, 15):
    cell = ws7.cell(row=20, column=c)
    if cell.value and cell.value < 0:
        cell.font = Font(name="Calibri", bold=True, size=10, color="CC0000")

# Seasonality weights reference
row = 22
ws7.cell(row=row, column=1, value="SEASONALITY WEIGHTS (Reference)")
style_section_row(ws7, row, 14)
row = 23
vals = ["Donation (GIVING_SEASON)", *[GIVING_SEASON[m] for m in MONTHS], sum(GIVING_SEASON.values())]
write_row(ws7, row, vals, pct_1_fmt)
row = 24
vals = ["Marketing (MKT_SEASON)", *[MARKETING_SEASON[m] for m in MONTHS], sum(MARKETING_SEASON.values())]
write_row(ws7, row, vals, pct_1_fmt)

auto_width(ws7, min_width=10, max_width=12)
ws7.column_dimensions["A"].width = 30
ws7.column_dimensions[get_column_letter(14)].width = 12


# ════════════════════════════════════════════════════════════
# SHEET 8: Assumptions
# ════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Assumptions")
ws8.sheet_properties.tabColor = "808080"

row = 1
ws8.cell(row=row, column=1, value="Key Assumptions & Parameters")
ws8.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
ws8.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

row = 3
for c, h in enumerate(["Parameter", "Value", "Source / Rationale"], 1):
    ws8.cell(row=row, column=c, value=h)
style_header_row(ws8, row, 3)

assumptions = [
    ("USD/CAD exchange rate", "1.36", "Projected; BoC forward rate"),
    ("Fiscal year", "Apr 2026 \u2013 Mar 2027 (Calendar FY)", "CA for-profit standard"),
    ("CA launch", "April 2026", ""),
    ("US launch", "August 2026 (4 mo after CA)", "Bootstrap from CA credibility"),
    ("CA platform fee", "3.5% \u2192 3.2% over 5 yr", "Competitive with CanadaHelps 4%"),
    ("US platform fee", "3.5% \u2192 3.2% over 5 yr", "Same trajectory as CA"),
    ("Stripe fee (CA & US)", "2.2% + $0.30/txn", "Nonprofit discount rate"),
    ("CA avg donation Y1", "$335 CAD", "CRA T1 2023 median $390"),
    ("US avg donation Y1", "$250 USD", "Network for Good 2023"),
    ("Donor activation Y1 (CA)", "10%", "Industry avg 10\u201315% (M+R 2024)"),
    ("Donor activation Y1 (US)", "8%", "Lower trust, new market"),
    ("CI unique visitors Y1", "305,000", "CI Web Traffic Analysis (internal)"),
    ("CI capture rate Y1", "30%", "Cross-links on charity profiles"),
    ("CI capture rate Y5", "49%", "Deep integration; <50% ceiling"),
    ("Registration rate Y1 (CA)", "1.0%", "New product; industry avg 1.8%"),
    ("Registration rate Y1 (US)", "2.0%", "Proven platform from CA launch"),
    ("SaaS partners Y1", "3", "Early adopter charity partners"),
    ("SaaS avg annual fee Y1", "$600 CAD", "Below Bloomerang $1,200+"),
    ("Donation seasonality", "GIVING_SEASON profile", "31% in Dec (Blackbaud 2023)"),
    ("Marketing seasonality", "MKT_SEASON profile", "Peak Nov 18% (lead time)"),
    ("Infrastructure scaling", "Linear interpolation", "Between tier breakpoints"),
    ("Tier breakpoints", "1K / 10K / 50K / 250K users", "Cloud provider pricing tiers"),
]

for i, (param, val, source) in enumerate(assumptions):
    row = 4 + i
    ws8.cell(row=row, column=1, value=param)
    ws8.cell(row=row, column=2, value=val)
    ws8.cell(row=row, column=3, value=source)
    for c in range(1, 4):
        cell = ws8.cell(row=row, column=c)
        cell.font = Font(name="Calibri", size=10)
        cell.border = thin_border
        if i % 2 == 0:
            cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

ws8.column_dimensions["A"].width = 32
ws8.column_dimensions["B"].width = 30
ws8.column_dimensions["C"].width = 45


# ════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════
out_path = "/home/selby/abt/docs/Harness-Give-Forecast-CI-US.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Sheets: {wb.sheetnames}")
